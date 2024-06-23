# backend/manual.py

import os
import requests
from io import BytesIO
import openpyxl
import psycopg2
from psycopg2.extras import execute_values
from langchain_openai import OpenAIEmbeddings
from langchain.text_splitter import CharacterTextSplitter
import numpy as np
from dotenv import load_dotenv

load_dotenv()

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
S3_DB_URL = os.getenv("S3_DB_URL", "http://s3_db:9000")
CHUNK_SIZE = int(os.getenv("CHUNK_SIZE"))
CHUNK_OVERLAP = int(os.getenv("CHUNK_OVERLAP"))
SEPARATOR = os.getenv("SEPARATOR")

embeddings = OpenAIEmbeddings(
    model="text-embedding-3-large",
    openai_api_key=OPENAI_API_KEY,
)

def normalize_vector(vector):
    norm = np.linalg.norm(vector)
    if norm == 0:
        return vector
    return vector / norm

def get_xlsx_files_from_s3():
    response = requests.get(f"{S3_DB_URL}/data/xlsx")
    if response.status_code == 200:
        return [file for file in response.json() if file.endswith('.xlsx')]
    else:
        print(f"Error fetching XLSX files: {response.status_code}")
        return []

def process_xlsx_file(file_name):
    response = requests.get(f"{S3_DB_URL}/data/xlsx/{file_name}")
    if response.status_code != 200:
        print(f"Error fetching file {file_name}: {response.status_code}")
        return []

    workbook = openpyxl.load_workbook(filename=BytesIO(response.content), read_only=True)

    file_data = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet_content = ""
        for row in sheet.iter_rows(values_only=True):
            if any(cell for cell in row):
                sheet_content += ' '.join(str(cell) if cell is not None else '' for cell in row) + "\n"

        text_splitter = CharacterTextSplitter(
            chunk_size=CHUNK_SIZE,
            chunk_overlap=CHUNK_OVERLAP,
            separator=SEPARATOR
        )
        chunks = text_splitter.split_text(sheet_content)

        for chunk in chunks:
            file_data.append({
                'file_name': file_name,
                'sheet_name': sheet_name,
                'manual': chunk
            })

    return file_data

def vectorize_and_normalize_text(text):
    try:
        vector = embeddings.embed_query(text)
        return normalize_vector(vector).tolist()
    except Exception as e:
        print(f"Error vectorizing text: {e}")
        return None

def create_table_if_not_exists(conn):
    with conn.cursor() as cur:
        cur.execute("""
        CREATE TABLE IF NOT EXISTS pgvector_manual (
            id SERIAL PRIMARY KEY,
            file_name TEXT,
            sheet_name TEXT,
            manual TEXT,
            manual_vector vector(3072)
        )
        """)
    conn.commit()

def save_to_database(conn, data):
    create_table_if_not_exists(conn)

    with conn.cursor() as cur:
        insert_query = """
        INSERT INTO pgvector_manual (file_name, sheet_name, manual, manual_vector)
        VALUES %s
        ON CONFLICT (file_name, sheet_name, manual) DO NOTHING
        """
        values = [
            (item['file_name'], item['sheet_name'], item['manual'], vectorize_and_normalize_text(item['manual']))
            for item in data
            if vectorize_and_normalize_text(item['manual']) is not None
        ]
        execute_values(cur, insert_query, values)

    conn.commit()

def main():
    db_connection_string = "postgresql://user:password@pgvector_manual:5432/manualdb"

    xlsx_files = get_xlsx_files_from_s3()

    conn = psycopg2.connect(db_connection_string)

    try:
        for file_name in xlsx_files:
            print(f"Processing file: {file_name}")
            file_data = process_xlsx_file(file_name)
            save_to_database(conn, file_data)
            print(f"Processed and saved data from {file_name}")

    except Exception as e:
        print(f"An error occurred: {e}")

    finally:
        conn.close()

if __name__ == "__main__":
    main()
